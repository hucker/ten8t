"""

The BaseCase.xlsx file contains:
  A           B
1 Status      Complete
2 Beginning   Yes
3 Middle	  Yes
4 Complete	  No
5 Post Mortem No


The BadCase.xlsx file contains 2pass/2fail/4skip (if skip on none):
The BadCase.xlsx file contains 2pass/6fail (if fail on none):

Status      Complete
Beginning	Yes
Middle	    Yes
Complete	No
Post Mortem	No
Missing

            Yes
            No
"""

import pandas as pd
import pytest
from openpyxl import load_workbook

from src import ten8t as t8

testdata = [

    # Testing cases for 'Sheet1' and specified columns
    ('Sheet1', 'A', 'B', 2, True, 'Beginning-Passed'),
    ('Sheet1', 'A', 'B', 3, True, 'Middle-Passed'),
    ('Sheet1', 'A', 'B', 4, False, 'Complete-Failed'),
    ('Sheet1', 'A', 'B', 5, False, 'Post Mortem-Failed'),

    # This gets the same pass/fail status, but there is no description
    # because using None for the desc-col has no default
    (None, None, None, 2, True, '-Passed'),
    (None, None, None, 3, True, '-Passed'),
    (None, None, None, 4, False, '-Failed'),
    (None, None, None, 5, False, '-Failed'),
]


@pytest.mark.parametrize('sheet, desc_col, val_col, row_start, expected_status, expected_msg', testdata)
def test_row_col_pass_fail_with_sheet(sheet, desc_col, val_col, row_start, expected_status, expected_msg):
    @t8.attributes(tag='foo')
    def check_pass_fail():
        wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
        yield from t8.rule_xlsx_a1_pass_fail(wb, sheet_name=sheet, val_col=val_col, row_start=str(row_start),
                                             desc_col=desc_col)

    s_func = t8.Ten8tFunction(check_pass_fail)
    ch = t8.Ten8tChecker(check_functions=[s_func], auto_setup=True)
    results = ch.run_all()

    # Verify that we got the status correct and that the messages are correct
    assert results[0].status is expected_status
    assert results[0].msg_rendered == expected_msg


@pytest.mark.parametrize('sheet, desc_col, val_col, row_start, expected_status, expected_msg', testdata)
def test_row_col_pass_fail_with_no_sheet(sheet, desc_col, val_col, row_start, expected_status, expected_msg):
    @t8.attributes(tag='foo')
    def check_pass_fail():
        wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
        yield from t8.rule_xlsx_a1_pass_fail(wb, val_col=val_col, row_start=str(row_start), desc_col=desc_col)

    s_func = t8.Ten8tFunction(check_pass_fail)
    ch = t8.Ten8tChecker(check_functions=[s_func], auto_setup=True)
    results = ch.run_all()

    # Verify that we got the status correct and that the messages are correct
    assert results[0].status is expected_status
    assert results[0].msg_rendered == expected_msg


@pytest.mark.parametrize('sheet, desc_col, val_col, row, expected_status, expected_msg', testdata)
def test_row_col_pass_fail_with_sheet_from_env(sheet, desc_col, val_col, row, expected_status, expected_msg):
    @t8.attributes(tag='foo')
    def check_pass_fail(wb):
        yield from t8.rule_xlsx_a1_pass_fail(wb, sheet_name=sheet, val_col=val_col, row_start=str(row),
                                             desc_col=desc_col)

    # Test that we can load workbooks from the environment.  Same test as above.
    s_func = t8.Ten8tFunction(check_pass_fail)
    ch = t8.Ten8tChecker(check_functions=[s_func],
                         env={"wb": load_workbook('./rule_xlsx/BaseCase.xlsx')},
                         auto_setup=True)
    results = ch.run_all()

    # Verify that we got the status correct and that the messages are correct
    assert results[0].status is expected_status
    assert results[0].msg_rendered == expected_msg


def test_row_col_pass_fail_with_auto_detect():
    @t8.attributes(tag='foo')
    def check_pass_fail(wb):
        yield from t8.rule_xlsx_a1_pass_fail(wb, sheet_name="Sheet1", val_col='B', row_start='2', row_end="auto",
                                             desc_col='A')

    s_func = t8.Ten8tFunction(check_pass_fail)
    wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
    ch = t8.Ten8tChecker(check_functions=[s_func], env={'wb': wb}, auto_setup=True)
    results = ch.run_all()

    assert len(results) == 4
    assert results[0].status is True
    assert results[1].status is True
    assert results[2].status is False
    assert results[3].status is False


def test_row_col_pass_fail_bad_sheet():
    def check_pass_fail(wb):
        yield from t8.rule_xlsx_a1_pass_fail(wb, sheet_name="Sheet666", val_col='B', row_start='2', row_end="auto",
                                             desc_col='A', first_if_missing=True)

    try:
        s_func = t8.Ten8tFunction(check_pass_fail)
        wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
        ch = t8.Ten8tChecker(check_functions=[s_func], env={'wb': wb}, auto_setup=True)
        _ = ch.run_all()
    except Exception as e:
        assert e



def test_row_col_pass_fail_with_hardcoded():
    @t8.attributes(tag='foo')
    def check_pass_fail(wb):
        yield from t8.rule_xlsx_a1_pass_fail(wb, sheet_name="Sheet1", val_col='B', row_start='2', row_end="5",
                                             desc_col='A')

    s_func = t8.Ten8tFunction(check_pass_fail)
    wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
    ch = t8.Ten8tChecker(check_functions=[s_func], env={'wb': wb}, auto_setup=True)
    results = ch.run_all()

    assert len(results) == 4
    assert results[0].status is True
    assert results[1].status is True
    assert results[2].status is False
    assert results[3].status is False


def test_row_col_df():
    @t8.attributes(tag='foo')
    def check_pass_fail(df):
        yield from t8.rule_xlsx_df_pass_fail(df, val_col='Complete', desc_col='Status')

    s_func = t8.Ten8tFunction(check_pass_fail)
    df = pd.read_excel('./rule_xlsx/BaseCase.xlsx', sheet_name='Sheet1')
    ch = t8.Ten8tChecker(check_functions=[s_func], env={'df': df}, auto_setup=True)
    results = ch.run_all()

    assert len(results) == 4
    assert sum(r.status for r in results) == 2
    assert results[0].status is True
    assert results[1].status is True
    assert results[2].status is False
    assert results[3].status is False


def test_bad_skip_on_null_df():
    @t8.attributes(tag='foo')
    def check_pass_fail(df):
        yield from t8.rule_xlsx_df_pass_fail(df, val_col='Complete', desc_col='Status', skip_on_none=True)

    s_func = t8.Ten8tFunction(check_pass_fail)
    df = pd.read_excel('./rule_xlsx/BadCase.xlsx', sheet_name='Sheet1')
    ch = t8.Ten8tChecker(check_functions=[s_func], env={'df': df}, auto_setup=True)
    results = ch.run_all()

    assert len(results) == 8
    assert sum(r.skipped for r in results if r.skipped) == 4


def test_bad_fail_on_null_df():
    @t8.attributes(tag='foo')
    def check_pass_fail(df):
        yield from t8.rule_xlsx_df_pass_fail(df, val_col='Complete', desc_col='Status', skip_on_none=False)

    s_func = t8.Ten8tFunction(check_pass_fail)
    df = pd.read_excel('./rule_xlsx/BadCase.xlsx', sheet_name='Sheet1')
    ch = t8.Ten8tChecker(check_functions=[s_func], env={'df': df}, auto_setup=True)
    results = ch.run_all()

    assert len(results) == 8
    assert sum(r.skipped for r in results if r.skipped) == 0
    assert sum(r.status is False for r in results if not r.status) == 6
